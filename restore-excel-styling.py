#!/usr/bin/env python3
"""
Restore Excel file styling and column widths from saved JSON.
Run this AFTER editing data.xlsx to restore the styling.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json
from pathlib import Path
from openpyxl.utils import get_column_letter

def create_style_from_dict(style_dict):
    """Create openpyxl style objects from a dictionary."""
    style = {}

    # Font styling
    if any(k in style_dict for k in ['font_name', 'font_size', 'font_bold', 'font_italic', 'font_color']):
        font_color = style_dict.get('font_color')
        style['font'] = Font(
            name=style_dict.get('font_name', 'Calibri'),
            size=style_dict.get('font_size', 11),
            bold=style_dict.get('font_bold', False),
            italic=style_dict.get('font_italic', False),
            color=font_color[2:] if font_color and font_color.startswith('FF') else font_color
        )

    # Fill styling
    if 'fill' in style_dict:
        fill_color = style_dict['fill']
        if fill_color and fill_color.startswith('FF'):
            style['fill'] = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    # Alignment
    if 'alignment_horizontal' in style_dict or 'alignment_vertical' in style_dict or 'alignment_wrap_text' in style_dict:
        style['alignment'] = Alignment(
            horizontal=style_dict.get('alignment_horizontal'),
            vertical=style_dict.get('alignment_vertical'),
            wrap_text=style_dict.get('alignment_wrap_text', False)
        )

    # Border
    if 'border' in style_dict:
        border_style = style_dict['border']
        borders = {}
        for side in ['left', 'right', 'top', 'bottom']:
            if side in border_style:
                borders[side] = Side(style=border_style[side])
        if borders:
            style['border'] = Border(**borders)

    # Number format
    if 'number_format' in style_dict:
        style['number_format'] = style_dict['number_format']

    return style

def restore_styling(excel_path, styling_path=None):
    """Restore Excel styling from a JSON file."""
    if styling_path is None:
        styling_path = Path(excel_path).parent / '.excel-styling.json'

    if not Path(styling_path).exists():
        print(f"❌ Styling file not found: {styling_path}")
        return False

    with open(styling_path, 'r') as f:
        styling_data = json.load(f)

    wb = load_workbook(excel_path)

    for sheet_name in wb.sheetnames:
        if sheet_name not in styling_data:
            print(f"⚠️  No styling data for sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        sheet_styling = styling_data[sheet_name]

        # Restore column widths
        for col_letter, width in sheet_styling.get('column_widths', {}).items():
            ws.column_dimensions[col_letter].width = width

        # Restore row heights
        for row_num_str, height in sheet_styling.get('row_heights', {}).items():
            row_num = int(row_num_str)
            ws.row_dimensions[row_num].height = height

        # Restore header styles
        for cell_coord, cell_style in sheet_styling.get('header_styles', {}).items():
            cell = ws[cell_coord]
            styles = create_style_from_dict(cell_style)
            for style_name, style_obj in styles.items():
                setattr(cell, style_name, style_obj)

    wb.save(excel_path)
    print(f"✅ Styling restored to: {excel_path}")
    return True

if __name__ == '__main__':
    import sys
    excel_file = sys.argv[1] if len(sys.argv) > 1 else 'data.xlsx'
    restore_styling(excel_file)
