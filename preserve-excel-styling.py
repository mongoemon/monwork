#!/usr/bin/env python3
"""
Preserve Excel file styling and column widths.
Run this BEFORE making changes to data.xlsx to save the styling info.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json
from pathlib import Path

def extract_cell_style(cell):
    """Extract styling from a cell."""
    style = {}

    if cell.fill:
        if cell.fill.start_color and cell.fill.start_color.rgb:
            style['fill'] = str(cell.fill.start_color.rgb)

    if cell.font:
        style['font_name'] = cell.font.name
        style['font_size'] = cell.font.size
        style['font_bold'] = cell.font.bold
        style['font_italic'] = cell.font.italic
        style['font_color'] = str(cell.font.color.rgb) if cell.font.color and cell.font.color.rgb else None

    if cell.alignment:
        style['alignment_horizontal'] = cell.alignment.horizontal
        style['alignment_vertical'] = cell.alignment.vertical
        style['alignment_wrap_text'] = cell.alignment.wrap_text

    if cell.border:
        border_dict = {}
        if cell.border.left:
            border_dict['left'] = cell.border.left.style
        if cell.border.right:
            border_dict['right'] = cell.border.right.style
        if cell.border.top:
            border_dict['top'] = cell.border.top.style
        if cell.border.bottom:
            border_dict['bottom'] = cell.border.bottom.style
        if border_dict:
            style['border'] = border_dict

    if cell.number_format and cell.number_format != 'General':
        style['number_format'] = cell.number_format

    return {k: v for k, v in style.items() if v is not None}

def save_styling(excel_path, output_path=None):
    """Save Excel styling information to a JSON file."""
    if output_path is None:
        output_path = Path(excel_path).parent / '.excel-styling.json'

    wb = load_workbook(excel_path)
    styling_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        styling_data[sheet_name] = {
            'column_widths': {},
            'row_heights': {},
            'header_styles': {},
            'cell_styles': {}
        }

        # Save column widths
        for col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            if width:
                styling_data[sheet_name]['column_widths'][col_letter] = width

        # Save row heights
        for row_num in ws.row_dimensions:
            height = ws.row_dimensions[row_num].height
            if height:
                styling_data[sheet_name]['row_heights'][row_num] = height

        # Save header row (row 1) styling
        for col in ws.iter_cols(min_row=1, max_row=1, values_only=False):
            for cell in col:
                if cell.value is not None:
                    cell_style = extract_cell_style(cell)
                    if cell_style:
                        styling_data[sheet_name]['header_styles'][cell.coordinate] = cell_style

    with open(output_path, 'w') as f:
        json.dump(styling_data, f, indent=2)

    print(f"✅ Styling saved to: {output_path}")
    return str(output_path)

if __name__ == '__main__':
    import sys
    excel_file = sys.argv[1] if len(sys.argv) > 1 else 'data.xlsx'
    save_styling(excel_file)
